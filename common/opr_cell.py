
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import inspect

MAXROW = None

def dprint(msg):
    frame_info = inspect.stack()[1]
    file_name = frame_info[1]
    line_number = frame_info[2]
    function_name = frame_info[3]
    print_msg = f"[print>>>>> line:{line_number} fuction:{function_name}] {msg}"
    # print_msg = f"[print]line_{line_number}_fuction_{function_name}_{file_name}] {msg}"
    dprint(print_msg)



"""按sheet名字cell写"""
def write_to_excel_cell(file, sheetname, row, column, value):
    work_sheet = file[sheetname]
    work_sheet.cell(row,column,value)  #修改
    # file.save(path) #保存

"""按sheetname cell读，如果读到公式，将返回公式表达式"""
def read_from_excel_cell(file, sheetname, row, column):
    work_sheet = file[sheetname]
    ret = work_sheet.cell(row, column).value
    # file.save(path) #保存
    return ret


"""如果读到公式，则返回为公式的值, 而不是公式的表达式"""
def data_only_read_from_excel_cell(file, sheetname, row, column):
    file = openpyxl.load_workbook(file, data_only=True)   #加载
    # file = xlrd.open_workbook(path)   #加载
    work_sheet = file[sheetname]
    ret = work_sheet.cell(row, column).value
    # file.save(path) #保存
    return ret


"""字符串查找  查找该字符串的单元格位置和内容"""
def find_string_in_excel(file, sheetname, target_string, result_row_list, result_column_list):
    sheet = file[sheetname]
    for row in sheet.iter_rows():
        for cell in row:
            if target_string in str(cell.value):
                # print("行：%d 列：%d" %(cell.row, cell.column))
                result_row_list.append(cell.row)
                result_column_list.append(cell.column)
            # else:
                # print("查无此人：%s",target_string)

"""字符串查找  查找该字符串的单元格位置和内容"""
# def find_string_in_excel(file, sheetname, target_string, result_row_list, result_column_list):
#     sheet = file[sheetname]
#     for row in sheet.iter_rows():
#         for cell in row:
#             if cell.value == target_string:
#                 print("行：%d 列：%d" %(cell.row, cell.column))
#                 result_row_list.append(cell.row)
#                 result_column_list.append(cell.column)



"""解冻单元格，解除筛选"""
def unfreeze_and_unfilt_cell(wb, sheetname):
    work_sheet = wb[sheetname]
    work_sheet.freeze_panes = None
    work_sheet.auto_filter.ref = None


def merge_cells_value(wb, sheetname=None, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None):
    """
        将靠上靠左的单元格的value赋值给合并后的单元格
    """
    sheet = wb[sheetname]
    if start_row > end_row:
        fmt = "{end_row} must be greater than {start_row}"
        raise ValueError(fmt.format(start_row=start_row, end_row=end_row))
    if start_column > end_column:
        fmt = "{end_column} must be greater than {start_column}"
        raise ValueError(fmt.format(start_column=start_column, end_column=end_column))
    if range_string is None:
        fmt = '{start_column}{start_row}:{end_column}{end_row}'
        range_string = fmt.format(start_row=start_row,
                                  start_column=get_column_letter(start_column),
                                  end_row=end_row,
                                  end_column=get_column_letter(end_column))

    v = None
    for cells in sheet[range_string]:
        for cell in cells:
            if cell.value is not None:
                v = cell.value
                break
        if v is not None:
            break
    # print(range_string)
    sheet.merge_cells(range_string=range_string)
    sheet['{0}{1}'.format(get_column_letter(start_column), str(start_row))] = v
# 使用方式


def unmerge_cells_value(wb, sheetname=None, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None):
    """
        将靠上靠左的单元格的value赋值给合并后的单元格
    """
    sheet = wb[sheetname]
    if start_row > end_row:
        fmt = "{end_row} must be greater than {start_row}"
        raise ValueError(fmt.format(start_row=start_row, end_row=end_row))
    if start_column > end_column:
        fmt = "{end_column} must be greater than {start_column}"
        raise ValueError(fmt.format(start_column=start_column, end_column=end_column))
    if range_string is None:
        fmt = '{start_column}{start_row}:{end_column}{end_row}'
        range_string = fmt.format(start_row=start_row,
                                  start_column=get_column_letter(start_column),
                                  end_row=end_row,
                                  end_column=get_column_letter(end_column))
    # print(range_string)
    sheet.unmerge_cells(range_string=range_string)









