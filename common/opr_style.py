import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl import Workbook
from common.common import *



def get_cell_color(wb, sheetname, row, col):
    work_sheet = wb[sheetname]
    fill = work_sheet.cell(row=row, column=col).fill
    hex_color = fill.start_color.index if fill.start_color else None
    return hex_color

def set_cell_color(wb, sheetname,row, col, color):
    worksheet = wb[sheetname]
    fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="DAEEF3")
    worksheet.cell(row=row, column=col).fill = fill


def set_row_color(wb, sheetname, target_row, col_start, col_end, color):
    worksheet = wb[sheetname]   
    fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    for i in (col_end - col_start + 1):
        worksheet.cell(row=target_row, column=col_start+i).fill = fill


def set_col_color(wb, sheetname, target_col, row_start, row_end, color):
    worksheet = wb[sheetname]   
    fill = PatternFill(start_color="darkGrid", end_color="darkGrid", fill_type=color)
    for i in (row_end - row_start + 1):
        worksheet.cell(row=row_start+i, column=target_col).fill = fill


def set_color_of_sheet(wb, sheetname):
    color = get_cell_color(wb, sheetname, 5, 3)
    print(color)
    set_col_color(wb, sheetname, 13, 2, 7, color)












