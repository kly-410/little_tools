#!/usr/bin/env python3
import openpyxl
import datetime
import xlrd
from openpyxl import load_workbook

# """
# 按sheet number 写
# """
# def write_to_excel_cell_num(path, sheetnumber, row, column, value):
#     file = openpyxl.load_workbook(path)   #加载
#     sheetnumber = file.active   #激活
#     sheetnumber.cell(row,column,value)  #修改
#     file.save(path) #保存


"""按sheet名字cell写"""
def write_to_excel_cell(file, sheetname, row, column, value):
    work_sheet = file[sheetname]
    work_sheet.cell(row,column,value)  #修改
    # file.save(path) #保存

""" row write """
def row_write_to_excel(file, sheetname, row, column_start, rowlist):
    for i in range(len(rowlist)):
        write_to_excel_cell(file, sheetname, row, column_start + i, rowlist[i])


""" column write """
def column_write_to_excel(file, sheetname, column, row_start, columnlist):
    for i in range(len(columnlist)):
        # print(columnlist[i])
        write_to_excel_cell(file, sheetname, row_start + i, column, columnlist[i])



"""按sheetname cell读，如果读到公式，将返回公式表达式"""
def read_from_excel_cell(file, sheetname, row, column):
    work_sheet = file[sheetname]
    ret = work_sheet.cell(row, column).value
    # file.save(path) #保存
    return ret


"""如果读到公式，则返回为公式的值, 而不是公式的表达式"""
def data_only_read_from_excel_cell(file, sheetname, row, column):
    file = openpyxl.load_workbook(file, data_only=True)   #加载
    # file = xlrd.open_workbook(path)   #加载
    work_sheet = file[sheetname]
    ret = work_sheet.cell(row, column).value
    # file.save(path) #保存
    return ret

""" row read """
def row_read_from_excel(file, sheetname, row, column_start, num, row_val_list):
    for i in range(num):
        row_val_list.append(read_from_excel_cell(file, sheetname, row, column_start + i))
        # print(row_val_list[i])


""" column read """
def column_read_from_excel(file, sheetname, column, row_start, num, column_val_list):
    for i in range(num):
        column_val_list.append(read_from_excel_cell(file, sheetname, row_start + i, column))
        # print(column_val_list[i])




""" row swag"""
def row_swag(file, sheetname, row0, row1,column_start, num):
    row_list_0 = []
    row_list_1 = []
    row_read_from_excel(file, sheetname, row0, column_start, num, row_list_0)
    row_read_from_excel(file, sheetname, row1, column_start, num, row_list_1)

    for i in range(num):
        tmp = row_list_0[i]
        row_list_0[i] = row_list_1[i]
        row_list_1[i] = tmp

    row_write_to_excel(file, sheetname, row0, column_start, row_list_0)
    row_write_to_excel(file, sheetname, row1, column_start, row_list_1)


""" column swag"""
def column_swag(file, sheetname, column0, column1, row_start, num):
    column_list_0 = []
    column_list_1 = []
    column_read_from_excel(file, sheetname, column0, row_start, num, column_list_0)
    column_read_from_excel(file, sheetname, column1, row_start, num, column_list_1)

    for i in range(num):
        tmp = column_list_0[i]
        column_list_0[i] = column_list_1[i]
        column_list_1[i] = tmp

    column_write_to_excel(file, sheetname, column0, row_start, column_list_0)
    column_write_to_excel(file, sheetname, column1, row_start, column_list_1)





""" row copy"""
def row_copy(file, sheetname, row_source, row_target,column_start, num):
    row_list_tmp = []
    row_read_from_excel(file, sheetname, row_source, column_start, num, row_list_tmp)
    row_write_to_excel(file, sheetname, row_target, column_start, row_list_tmp)


""" column copy"""
def column_copy(file, sheetname, column_source, column_target, row_start, num):
    column_list_tmp = []
    column_read_from_excel(file, sheetname, column_source, row_start, num, column_list_tmp)
    column_write_to_excel(file, sheetname, column_target, row_start, column_list_tmp)


def del_row_in_excel(file, sheetname, to_del_row_list):
    work_sheet = file[sheetname]
    for i in range(len(to_del_row_list)):
        work_sheet.delete_rows((to_del_row_list[i]-i)) #删除后下边的单元格上移动


def del_column_in_excel(file, sheetname, to_del_column_list):
    work_sheet = file[sheetname]
    for i in range(len(to_del_column_list)):
        work_sheet.delete_cols((to_del_column_list[i]-i)) #删除后下边的单元格上移动


"""字符串查找  查找该字符串的单元格位置和内容"""
# def find_string_in_excel(file, sheetname, target_string, result_row_list, result_column_list):
#     sheet = file[sheetname]
#     for row in sheet.iter_rows():
#         for cell in row:
#             if cell.value == target_string:
#                 print("行：%d 列：%d" %(cell.row, cell.column))
#                 result_row_list.append(cell.row)
#                 result_column_list.append(cell.column)

def find_string_in_excel(file, sheetname, target_string, result_row_list, result_column_list):
    sheet = file[sheetname]
    for row in sheet.iter_rows():
        for cell in row:
            if target_string in str(cell.value):
                # print("行：%d 列：%d" %(cell.row, cell.column))
                result_row_list.append(cell.row)
                result_column_list.append(cell.column)
            # else:
            #     print("查无此人：%s",target_string)




class SysTime:
    def __init__(self):
        self.now = datetime.datetime.now()
        self.year = self.now.year
        self.month = self.now.month
        self.day = self.now.day
        self.hour = self.now.hour
        self.minute = self.now.minute


class SumExcel:
    def __init__(self, file, sheetname, target_c_or_r, start_of_target, num):
        self.file = file
        self.sheetname = sheetname
        self.target = target_c_or_r
        self.start = start_of_target
        self.num =num
        self.cachelist = []
        self.tmp = 0

    def return_column_sum(self):
        column_read_from_excel(self.file, self.sheetname, self.target, self.start, self.num, self.cachelist)
        for i in range(len(self.cachelist)):
            if self.cachelist[i] == None:
                self.cachelist[i] = 0  
            self.tmp += self.cachelist[i]
        print("第%d列, %d-%d行的和为：%d" % (self.target, self.start, self.num + self.start, self.tmp))
        return self.tmp

    def return_row_sum(self):
        row_read_from_excel(self.file, self.sheetname, self.target, self.start, self.num, self.cachelist)
        for i in range(len(self.cachelist)):
            if self.cachelist[i] == None:
                self.cachelist[i] = 0  
            self.tmp += self.cachelist[i]
        print("第%d行, %d-%d列的和为：%d" % (self.target, self.start, self.num + self.start, self.tmp))
        return self.tmp



class GetSearchInExcel:
    def __init__(self, wb, sheet_name, str ):
        self.wb = wb
        self.sheet_name = sheet_name
        self.str = str
        self.row_val_list = []
        self.column_val_list = []

    #找一个最大行，列的字符串,用于搜索    
    def get_max_row(self):
        find_string_in_excel(self.wb, self.sheet_name, self.str , self.row_val_list, self.column_val_list)
        return self.row_val_list[-1]

    def get_max_column(self):
        find_string_in_excel(self.wb, self.sheet_name, self.str , self.row_val_list, self.column_val_list)
        return self.column_val_list[-1]



def del_none_row_and_col(wb, sheet_name):
    work_sheet = wb[sheet_name]
    to_del_row_list = []
    to_del_col_list = []

    row_val_list = []
    column_val_list = []
    maxrow = GetSearchInExcel(wb, sheet_name, "合计")
    maxrow_num = maxrow.get_max_row()
    maxcol = GetSearchInExcel(wb, sheet_name, "1-12月合计")
    maxcol_num = maxcol.get_max_column() +3  #由表格获得,+3是多删除一点空白列
    print("最大列：", maxcol_num)


    find_string_in_excel(wb, sheet_name, "本月预算", row_val_list, column_val_list)
    print("基地址(本月预算)：行_%d,列_%d" %(row_val_list[0],column_val_list[0]))

    find_string_in_excel(wb, sheet_name, "累计费用", row_val_list, column_val_list)
    print("基地址(累计费用)：行_%d,列_%d" %(row_val_list[1],column_val_list[1]))


    #获取坐标
    base_row_start = row_val_list[0]
    base_column_start = column_val_list[0]
    base_row_end = row_val_list[1]
    base_column_end = column_val_list[1]

    if row_val_list[0] != row_val_list[1]:
        print("基地址开始行，获取错误！！！！")
        return
    
    column_size = base_column_end - base_column_start + 1 
    # print("开始行：%d, 最大行：%d" %(base_row_start, maxrow_num ))


    # print("************************************************")
    # print("****************开始删除表格********************")

    for row_num in range(base_row_start, maxrow_num):

        flag_all_none = 0
        for i in range(base_column_start, (base_column_end +1)):
            # cell = work_sheet.cell(row=row_num, column=i)
            tmp = work_sheet.cell(row_num, i).value
            # print(tmp)
            if tmp == None or tmp == 0:
                flag_all_none += 1
                
        if flag_all_none == column_size:
            to_del_row_list.append(row_num)
        # del_row_list.append(row_num)
            # tmp1 = work_sheet.cell(row, base_column_start - 1).value
            # if tmp1 == "小计":
            #     tmp2 = work_sheet.cell(row, base_column_start - 2).value
            #     if tmp2 != None:
            #         work_sheet.delete_rows(row - flag_has_been_del)
            #         flag_has_been_del += 1
            #     else:
            #         work_sheet.delete_rows(row - flag_has_been_del)
            #         flag_has_been_del += 1

            # elif tmp1 != "小计":
            #     tmp3 = work_sheet.cell(row, base_column_start - 2).value
            #     if tmp3 != None:
            #         row_copy(wb, sheet_name, row, row+1, (base_column_start - 2), 1)
            #         work_sheet.delete_rows(row - flag_has_been_del)
            #         flag_has_been_del += 1
            #     else:
            #         row_copy(wb, sheet_name, row, row+1, (base_column_start - 2), 1)
            #         work_sheet.delete_rows(row - flag_has_been_del)
            #         flag_has_been_del += 1                
    # print(to_del_row_list)
    del_row_in_excel(wb, sheet_name, to_del_row_list)



    del_col_start = base_column_end + 2
    del_col_end = maxcol_num 
    for i in range(del_col_start, del_col_end + 1):
        to_del_col_list.append(i)

    del_column_in_excel(wb, sheet_name, to_del_col_list) 
    # print("删除空白行，列成功！") 


def write_sum_to_xiaoji_row(wb, sheet_name):
    list_ABC= [ {'1': 'A', 
                 '2': 'B', 
                 '3': 'C',
                 '4': 'D',
                 '5': 'E',
                 '6': 'F',
                 '7': 'G',
                 '8': 'H',
                 '9': 'I'}]

    work_sheet = wb[sheet_name]

    row_base_addr_list = []
    column_base_addr_list = []
    xioaji_and_tushu_row_list = []

    maxrow = GetSearchInExcel(wb, sheet_name, "合计")
    maxrow_num = maxrow.get_max_row()

    find_string_in_excel(wb, sheet_name, "本月预算", row_base_addr_list, column_base_addr_list)
    # print("基地址(本月预算)：行_%d,列_%d" %(row_base_addr_list[0],column_base_addr_list[0]))
    col_base_start = column_base_addr_list[0]

    for column in range(col_base_start,col_base_start + 5):
        row_start = row_base_addr_list[0] +1
        row_end = None
        for i in range(row_base_addr_list[0] +1, maxrow_num):
            tmp1 = work_sheet.cell(i, column_base_addr_list[0] - 1).value
            if tmp1 == "小计":
                # print(i)
                row_end = i - 1
                value = "=sum(" + str(list_ABC[0][str(column)]) + str(row_start) + ":" + str(list_ABC[0][str(column)]) + str(row_end) +")"
                # print(value)
                work_sheet.cell(i,column,value)
                row_start = i + 1
            if tmp1 == "图书杂志费":
                row_start += 1

    #获取小记所在行,并完成每一列小计的合计
    for i in range(row_base_addr_list[0] +1, maxrow_num):
        tmp1 = work_sheet.cell(i, column_base_addr_list[0] - 1).value
        if tmp1 == "小计":
            xioaji_and_tushu_row_list.append(i)
        if tmp1 == "图书杂志费":
            xioaji_and_tushu_row_list.append(i)   
    # print(xioaji_and_tushu_row_list)

    #写入合计
    for column in range(col_base_start,col_base_start + 5 ):
        # print("column:",column)
        str_to_write = "="
        for i in range(len(xioaji_and_tushu_row_list)):
            str_to_write += "+" + str(list_ABC[0][str(column)]) + str(xioaji_and_tushu_row_list[i])
        # print(str_to_write)
        work_sheet.cell(maxrow_num,column,str_to_write)
    # print("合计成功！")


def copy_data_from_src(wb, sheet_name):
    work_sheet = wb[sheet_name]
    systime = SysTime()
    month = systime.month
    now = systime.now
    print("现在是%d月，进行%d月表格处理" %(month, (int(month) - 1)))

    maxrow = GetSearchInExcel(wb, sheet_name, "合计")
    maxrow_num = maxrow.get_max_row()

    row_base_addr_list = []
    column_base_addr_list = []
    month_row_list = []
    month_column_list = []

    row_of_month_real_money = 0
    col_of_month_real_money = 0
    row_of_month_plan_money = 0
    col_of_month_plan_money = 0

    month_str = (str(month - 1) + "月")

    find_string_in_excel(wb, sheet_name, "本月预算", row_base_addr_list, column_base_addr_list)
    find_string_in_excel(wb, sheet_name, month_str, month_row_list, month_column_list)
    n = 4
    for i in range(len(month_row_list)):
        tmp1 = work_sheet.cell(month_row_list[i], month_column_list[i]).value


        if tmp1 == (str(month - 1) + "月") and month_row_list[i] ==row_base_addr_list[0]:
            # print("%d月 ：行_%d,列_%d" %(month, month_row_list[i], month_column_list[i]))
            row_of_month_real_money = month_row_list[i]
            col_of_month_real_money = month_column_list[i]
            row_of_month_plan_money = month_row_list[i]
            col_of_month_plan_money = month_column_list[i]
            # print(month_row_list[i]+1)

            # print(maxrow_num - month_row_list[i]-1)
           
            column_copy(wb, sheet_name, col_of_month_real_money, n, month_row_list[i]+1, maxrow_num - month_row_list[i]-1)
            n = n-1

def have_processed_check1(file, sheet_name):
    wb = openpyxl.load_workbook(file,data_only=True)   #加载

    target_str_row = []
    target_str_row.append(None)
    target_str_col = [] 

    find_string_in_excel(wb, sheet_name, "1-12月合计", target_str_row, target_str_col)
    # print("处理标记行：",target_str_row[0])
    if target_str_row[0] == None:
        print("#已为：发送版，无需再次处理")
        # print("%s已完成：透视表 》》》》》底稿，无需再次处理",file)
        return -1
    wb.save(file)



def process_single_excel(file):
    sheet_name = "安环部预算执行表"
    # tmp = have_processed_check1(file, sheet_name)
    # if tmp == -1:
    #     return
    
    wb = openpyxl.load_workbook(file,data_only=True)   #加载
    copy_data_from_src(wb, sheet_name)
    del_none_row_and_col(wb, sheet_name)
    write_sum_to_xiaoji_row(wb, sheet_name)
    wb.save(file)

def print_log():
    print("                 _             _   ")
    print(" _   _ _   _ ___| |_ __ _ _ __| |_ ")
    print("| | | | | | / __| __/ _` | '__| __|")
    print("| |_| | |_| \__ \ || (_| | |  | |_ ")
    print(" \__,_|\__,_|___/\__\__,_|_|   \__|")









